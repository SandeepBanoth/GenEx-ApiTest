import openai
import openpyxl
import re
 
 
 
def test_excel_creator(swagger):
 
   deployment_name="EnggExGPT35"
   openai.api_type="azure"
   openai.api_key="0f13fd2e937642efb8a422394b85cb9d"
   openai.api_base="https://enggexopenai.openai.azure.com/"
   openai.api_version="2023-07-01-preview"
 
 
   api_details_in_excel="""
   Tool  Base URL API Endpoint   HTTP Method Request Header Request Body   File Location  Path Parameters   Query/header Parameters Authentication Type  Authentication Credentials Expected Status Code Response Headers  Assertions
   Rest Assured   https://reqres.in/api   /users/2 GET                        200      data.first_name: 'Janet'
   Rest Assured   https://reqres.in/api   /users   POST  ContentType.JSON  "{\n"+" \"name\": \"Sam\", \n" +" \"job\": \"Project Leader\"\n"+ "}"                  201      name: 'Sam', job: contains 'Leader'
   Rest Assured   https://reqres.in/api.users   /{id} GET            id=3           300      data.last_name: 'Wong'
   """
   output_format="""
   Test Case 1:
   - Tool= Rest Assured
   - Base URL= https://reqres.in/api
   - API Endpoint= /users/2
   - HTTP Method= GET
   - Request Header= None
   - Request Body= None
   - File Location= None
   - Path Parameters= None
   - Query/Header Parameters= None
   - Authentication Type= None
   - Authentication Credentials= None
   - Expected Status Code= 200
   - Response Headers= None
   - Assertions= data.first_name: 'Janet'
 
   Test Case 2:
   - Tool= Rest Assured
   - Base URL= https://reqres.in/api
   - API Endpoint= /users
   - HTTP Method= POST
   - Request Header= ContentType.JSON
   - Request Body= {"name": "Sam", "job": "Project Leader"}
   - File Location= None
   - Path Parameters= None
   - Query/Header Parameters= None
   - Authentication Type= None
   - Authentication Credentials= None
   - Expected Status Code= 201
   - Response Headers= None
   - Assertions= name: 'Sam', job: contains 'Leader', etc
 
   and so on..
   """
 
   prompt=f"""I have an API which is being tested using restassured for various purposes. I have provided you the
   {{{swagger}}} mentioned in triple delimeters.Consider you are an expert in api, api documentation, api testing, and test case and test scenario generation.
   I want to create an excel which contains all the details of api tesing using which I can do
   the api testing using restassured. I have provided you some examples {{{api_details_in_excel}}} for how the details in excel
   should look like. Your task is to read the swagger and give me the details to fill in excel for api testing with all possible details according to the given {{{output_format}}} only.
   Give me as many testcases as possible covering various aspects.
   The details you provide me are the following-
   Tool,
   Base URL,
   API Endpoint,
   HTTP Method,
   Request Header,
   Request Body,
   File Location,
   Path Parameters,  
   Query/header,
   Parameters,
   Authentication Type,
   Authentication Credentials,
   Expected Status Code,
   Response Headers,
   Assertions.
 
   Note: create as many testcases and scenarios/details as possible. Min 7 to 8 testcases must be given by you. Give the output strictly as per the {{{output_format}}} only. In output use Tool= Base URL= and so on, not Tool: Base URL: so on. Dont write any comments or explanation, just give the test cases. I will put each of them in a new row."""
 
 
   result=openai.ChatCompletion.create(
      messages=[{"role": "user", "content": prompt}],
      temperature=0,
      max_tokens= 10000,
      engine= deployment_name
   )
 
   chat_response=result.choices[0].message.content
   print(f'{chat_response}')
   #return chat_response
 
   fp = open("D:\GenAI API-testing\TestCases\Testcase.txt", 'w')
   fp.write(chat_response)
   fp.close()
 
 
   # Read the test cases from the text file
   with open('D:/GenAI API-testing/TestCases/Testcase.txt', 'r') as file:
      test_cases = file.read()
 
 
   # Split the test cases by 'Test Case' headers
   test_case_list = re.split(r'Test Case \d+:', test_cases)[1:]
 
   # Create a new Excel workbook and select the default sheet
   workbook = openpyxl.Workbook()
   sheet = workbook.active
 
   # Write the headings in the first row
   headings = [
      'Tool', 'Base URL', 'API Endpoint', 'HTTP Method', 'Request Header', 'Request Body',
      'File Location', 'Path Parameters', 'Query/header Parameters', 'Authentication Type',
      'Authentication Credentials', 'Expected Status Code', 'Response Headers', 'Assertions'
   ]
 
   for col, heading in enumerate(headings, start=1):
      sheet.cell(row=1, column=col, value=heading)
 
   # Write test case values to the worksheet
   for row, test_case in enumerate(test_case_list, start=2):
      details = [line.strip() for line in test_case.split('\n') if line.strip()]
      for col, detail in enumerate(details, start=1):
         # Extract the value part after ':'
         value = detail.split('= ')[-1].strip()
         sheet.cell(row=row, column=col, value=value)
 
   # Save the Excel file
   workbook.save('D:\GenAI API-testing\TestCases\Test_cases.xlsx')

